import json
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Side, Border, Alignment


def get_answers(question):
    # return " \n".join(question["answers"]["text"])
    return question["answers"]["text"]


def load_and_sort_data(file):
    data = json.load(file)["data"]
    # return sorted(data, key=lambda i: i['id'], reverse=True)
    return data


def load_and_sort_result(file):
    data = json.load(file)
    return list(data.items())
    # return [(k, data[k]) for k in sorted(data, key=data.get, reverse=True)]


def combine_results(data_path_eng, results_path_eng, data_path_slo_auto, results_path_slo_auto,
                    data_path_slo, results_path_slo, output_path):
    with open(data_path_eng, encoding='utf-8') as f:
        data_eng = load_and_sort_data(f)
    with open(data_path_slo_auto, encoding='utf-8') as f:
        data_slo_auto = load_and_sort_data(f)
    with open(data_path_slo, encoding='utf-8') as f:
        data_slo = load_and_sort_data(f)

    with open(results_path_eng, encoding='utf-8') as f:
        results_eng = load_and_sort_result(f)
    with open(results_path_slo_auto, encoding='utf-8') as f:
        results_slo_auto = load_and_sort_result(f)
    with open(results_path_slo, encoding='utf-8') as f:
        results_slo = load_and_sort_result(f)

    wb = Workbook()
    ws1 = wb.active
    # t_border = NamedStyle(name="border")
    # bd = Side(style='thick', color="000000")
    # t_border.border = Border(right=bd)
    top_alignment = Alignment(vertical="center")

    row_index = 1
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    ws1.merge_cells(start_row=1, start_column=4, end_row=1, end_column=6)
    ws1.merge_cells(start_row=1, start_column=7, end_row=1, end_column=9)
    # ws1.merge_cells('A1:C1')
    # ws1.merge_cells('D1:F1')
    # ws1.merge_cells('G1:H1')
    ws1.cell(column=1, row=row_index, value="English")
    ws1.cell(column=4, row=row_index, value="Computer Translation")
    ws1.cell(column=7, row=row_index, value="Human Translation")

    row_index += 1
    ws1.cell(column=1, row=row_index, value="Question")
    ws1.cell(column=2, row=row_index, value="Answers")
    ws1.cell(column=3, row=row_index, value="Prediction")
    ws1.cell(column=4, row=row_index, value="Question")
    ws1.cell(column=5, row=row_index, value="Answers")
    ws1.cell(column=6, row=row_index, value="Prediction")
    ws1.cell(column=7, row=row_index, value="Question")
    ws1.cell(column=8, row=row_index, value="Answers")
    ws1.cell(column=9, row=row_index, value="Prediction")

    context_displayed = []
    for i, result_eng_tuple in enumerate(results_eng):
        res_id = result_eng_tuple[0]

        result_eng = result_eng_tuple[1].strip()
        result_slo_auto = results_slo_auto[i][1].strip()
        result_slo = results_slo[i][1].strip()

        id_q = res_id.split("_")
        index_doc = int(id_q[0])
        index_paragraph = int(id_q[1])

        qas_eng = data_eng[i]
        qas_slo_auto = data_slo_auto[i]
        qas_slo = data_slo[i]

        if (index_doc, index_paragraph) not in context_displayed:
            context_displayed.append((index_doc, index_paragraph))
            context_eng = qas_eng["context"]
            context_slo_auto = qas_slo_auto["context"]
            context_slo = qas_slo["context"]
            row_index += 1
            ws1.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=3)
            ws1.merge_cells(start_row=row_index, start_column=4, end_row=row_index, end_column=6)
            ws1.merge_cells(start_row=row_index, start_column=7, end_row=row_index, end_column=9)
            ws1.cell(column=1, row=row_index, value=context_eng)
            ws1.cell(column=4, row=row_index, value=context_slo_auto)
            ws1.cell(column=7, row=row_index, value=context_slo)

        question_eng = qas_eng["question"]
        question_slo_auto = qas_slo_auto["question"]
        question_slo = qas_slo["question"]

        answers_eng = get_answers(qas_eng)
        answers_slo_auto = get_answers(qas_slo_auto)
        answers_slo = get_answers(qas_slo)

        row_index += 1
        if len(answers_eng) > 1:
            ws1.merge_cells(start_row=row_index, start_column=1, end_row=row_index + len(answers_eng) - 1, end_column=1)
            ws1.merge_cells(start_row=row_index, start_column=3, end_row=row_index + len(answers_eng) - 1, end_column=3)
            ws1.merge_cells(start_row=row_index, start_column=4, end_row=row_index + len(answers_eng) - 1, end_column=4)
            ws1.merge_cells(start_row=row_index, start_column=6, end_row=row_index + len(answers_eng) - 1, end_column=6)
            ws1.merge_cells(start_row=row_index, start_column=7, end_row=row_index + len(answers_eng) - 1, end_column=7)
            ws1.merge_cells(start_row=row_index, start_column=9, end_row=row_index + len(answers_eng) - 1, end_column=9)

        ws1.cell(column=1, row=row_index, value=question_eng).alignment = top_alignment
        ws1.cell(column=3, row=row_index, value=result_eng).alignment = top_alignment
        ws1.cell(column=4, row=row_index, value=question_slo_auto).alignment = top_alignment
        ws1.cell(column=6, row=row_index, value=result_slo_auto).alignment = top_alignment
        ws1.cell(column=7, row=row_index, value=question_slo).alignment = top_alignment
        ws1.cell(column=9, row=row_index, value=result_slo).alignment = top_alignment

        for a_i in range(len(answers_eng)):
            ws1.cell(column=2, row=row_index, value=answers_eng[a_i])
            ws1.cell(column=5, row=row_index, value=answers_slo_auto[a_i])
            ws1.cell(column=8, row=row_index, value=answers_slo[a_i])
            row_index += 1

    wb.save(filename=output_path)


combine_results("../data/translations/test_eng.json", "results/results_eng.json",
                "../data/translations/test_auto.json", "results/results_slo_auto.json",
                "../data/translations/test_slo.json", "results/results_slo.json",
                "output/combined_results.xlsx")
