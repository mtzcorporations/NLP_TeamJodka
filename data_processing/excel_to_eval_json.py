import json
import os
import glob
import openpyxl


def excel_to_eval_json(paths, output, auto_translation_json=None):
    if not isinstance(paths, list):
        if os.path.isdir(paths):
            paths = glob.glob(paths + "/*.xlsx")
        else:
            paths = [paths]

    json_eng = {"version": "v2.0", "language": "english", "data": []}
    json_slo = {"version": "v2.0", "language": "slovene", "data": []}

    if auto_translation_json is not None:
        with open(auto_translation_json, encoding="utf8") as f:
            auto_translation_json = json.load(f)["data"]
        json_auto = {"version": "v2.0", "language": "slovene", "data": []}

    print("Input files:")
    for path in paths:
        print(path)
        wookbook = openpyxl.load_workbook(path)
        worksheet = wookbook.active
        index = 0
        while index < worksheet.max_row:
            index += 1
            left = worksheet.cell(index, 1).value.strip()
            right = worksheet.cell(index, 2).value.strip()
            if "_" in left and "#" in left and left == right:
                context_id = left
                podatki = left.split("#")
                num_of_questions = int(len(podatki) / 2)
                num_of_answers = []
                for i, item in enumerate(podatki):
                    if i % 2 != 0:
                        item = item.split("-")
                        num_of_answers.append(len(item))
                index += 1
                context_eng = worksheet.cell(index, 1).value.strip()
                context_slo = worksheet.cell(index, 2).value.strip()

                if auto_translation_json is not None:
                    id_split = context_id.split("_")
                    document = int(id_split[0])
                    paragraph = int(id_split[1])
                    context_auto = auto_translation_json[document]["paragraphs"][paragraph]["context"]

                for i in range(num_of_questions):
                    answers_eng = {"text": [], "answer_start": []}
                    answers_slo = {"text": [], "answer_start": []}
                    answers_auto = {"text": [], "answer_start": []}
                    index += 1
                    question_eng = worksheet.cell(index, 1).value.strip()
                    question_slo = worksheet.cell(index, 2).value.strip()

                    if auto_translation_json is not None:
                        qas_split = id_split[2].split("#")
                        question = int(qas_split[2 * i])
                        question_auto = auto_translation_json[document]["paragraphs"][paragraph]["qas"][question][
                            "question"]

                    for j in range(num_of_answers[i]):
                        index += 1
                        answer_eng = worksheet.cell(index, 1).value.strip()
                        answer_slo = worksheet.cell(index, 2).value.strip()
                        index_eng = context_eng.find(answer_eng)
                        index_slo = context_slo.find(answer_slo)
                        if index_slo == -1:
                            print("ERROR in answer", answer_slo)

                        answers_eng["text"].append(answer_eng)
                        answers_slo["text"].append(answer_slo)
                        answers_eng["answer_start"].append(index_eng)
                        answers_slo["answer_start"].append(index_slo)

                        if auto_translation_json is not None:
                            answer_split = qas_split[2 * i + 1].split("-")
                            answer = int(answer_split[j])
                            answer_auto = auto_translation_json[document]["paragraphs"][paragraph]["qas"][question]["answers"][answer]
                            answers_auto["text"].append(answer_auto["text"])
                            answers_auto["answer_start"].append(answer_auto["answer_start"])


                    json_eng["data"].append(
                        {"id": f'{context_id}&{i}', "title": "", "context": context_eng, "question": question_eng,
                         "answers": answers_eng})
                    json_slo["data"].append(
                        {"id": f'{context_id}&{i}', "title": "", "context": context_slo, "question": question_slo,
                         "answers": answers_slo})
                    if auto_translation_json is not None:
                        json_auto["data"].append(
                            {"id": f'{context_id}&{i}', "title": "", "context": context_auto, "question": question_auto,
                             "answers": answers_auto})

    with open(output + "_eng.json", "w", encoding="utf8") as f:
        f.write(json.dumps(json_eng, ensure_ascii=False))

    with open(output + "_slo.json", "w", encoding="utf8") as f:
        f.write(json.dumps(json_slo, ensure_ascii=False))

    if auto_translation_json is not None:
        with open(output + "_auto.json", "w", encoding="utf8") as f:
            f.write(json.dumps(json_auto, ensure_ascii=False))


excel_to_eval_json("../data/translations", "../data/translations/test", "output/dev-v2.0_translated_corrected.json")
